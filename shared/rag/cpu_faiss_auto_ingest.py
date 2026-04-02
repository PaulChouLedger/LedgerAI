#!/usr/bin/env python3
"""
CPU FAISS Auto-Ingestion System
Monitors data/input/ for new guideline files and automatically updates embeddings
"""

import os
import time
import json
import pickle
import hashlib
from pathlib import Path
from typing import Dict, List, Any
import numpy as np
import faiss
from sentence_transformers import SentenceTransformer
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    try:
        import pandas as pd
        EXCEL_SUPPORT = True
    except ImportError:
        EXCEL_SUPPORT = False
        print("[Auto-Ingest] ⚠️ Excel support not available. Install openpyxl or pandas")

# PDF and DOCX support
try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("[Auto-Ingest] ⚠️ PDF support not available. Install PyPDF2")

try:
    import docx
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False
    print("[Auto-Ingest] ⚠️ DOCX support not available. Install python-docx")

class CPUFAISSAutoIngest:
    """Auto-ingestion system for CPU FAISS"""
    
    def __init__(self):
        # Use absolute paths based on container working directory (/app)
        # Data lives at /app/data (symlinked by native launch scripts)
        base_dir = Path("/app/data")
        self.input_dir = base_dir / "input"
        self.parsed_dir = base_dir / "parsed"  # For future use if needed
        self.cpu_embeddings_dir = base_dir / "embeddings"
        self.model_name = "all-distilroberta-v1"
        self.embedding_dimension = 768
        
        # Initialize components
        self.model = None
        self.index = None
        self.chunks = []
        self.metadata = []
        self.state_file = base_dir / "cpu_ingest_state.json"
        self.state = self._load_state()
        
        # Ensure directories exist
        self.input_dir.mkdir(parents=True, exist_ok=True)
        self.parsed_dir.mkdir(parents=True, exist_ok=True)
        self.cpu_embeddings_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize model
        self._load_model()
        
        # Initialize file watcher
        self.observer = None
        self.watching = False
    
    def _load_model(self):
        """Load the sentence transformer model"""
        try:
            self.model = SentenceTransformer(self.model_name)
            print(f"[Auto-Ingest] ✅ Loaded model: {self.model_name}")
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to load model: {e}")
            raise
    
    def _load_state(self) -> Dict:
        """Load ingestion state"""
        if self.state_file.exists():
            try:
                with open(self.state_file, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"[Auto-Ingest] ⚠️ Failed to load state: {e}")
        return {"processed_files": {}, "last_scan": None}
    
    def _save_state(self):
        """Save ingestion state"""
        try:
            with open(self.state_file, 'w') as f:
                json.dump(self.state, f, indent=2)
        except Exception as e:
            print(f"[Auto-Ingest] ⚠️ Failed to save state: {e}")
    
    def _get_file_hash(self, file_path: Path) -> str:
        """Get file hash for change detection"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception as e:
            print(f"[Auto-Ingest] ⚠️ Failed to get hash for {file_path}: {e}")
            return ""
    
    def chunk_text(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """
        Improved semantic-aware chunking that respects document structure
        
        Strategy:
        1. Split by paragraphs first (double newlines)
        2. Within paragraphs, respect sentence boundaries
        3. Maintain overlap between chunks for context preservation
        4. Ensure chunks are semantically coherent
        """
        import re
        
        # First, split by paragraphs (double newlines or section breaks)
        paragraphs = re.split(r'\n\s*\n+', text.strip())
        
        chunks = []
        current_chunk = []
        current_length = 0
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            para_words = para.split()
            para_length = len(para_words)
            
            # If paragraph fits in current chunk, add it
            if current_length + para_length <= chunk_size:
                current_chunk.extend(para_words)
                current_length += para_length
            else:
                # Save current chunk if it has content
                if current_chunk:
                    chunk_text = " ".join(current_chunk)
                    if chunk_text.strip():
                        chunks.append(chunk_text.strip())
                
                # Start new chunk with overlap
                # Take last 'overlap' words from previous chunk for context
                if chunks and overlap > 0:
                    prev_words = chunks[-1].split() if chunks else []
                    overlap_words = prev_words[-overlap:] if len(prev_words) >= overlap else prev_words
                    current_chunk = overlap_words + para_words
                    current_length = len(current_chunk)
                else:
                    current_chunk = para_words
                    current_length = para_length
            
            # If current chunk exceeds size, split it at sentence boundaries
            while current_length > chunk_size:
                # Try to split at sentence boundary within current chunk
                chunk_text = " ".join(current_chunk)
                sentences = re.split(r'([.!?]+\s+)', chunk_text)
                
                # Reconstruct sentences properly
                proper_sentences = []
                for i in range(0, len(sentences) - 1, 2):
                    if i + 1 < len(sentences):
                        proper_sentences.append(sentences[i] + sentences[i + 1])
                    else:
                        proper_sentences.append(sentences[i])
                
                # Find split point
                split_idx = 0
                accumulated_length = 0
                for sent in proper_sentences:
                    sent_words = sent.split()
                    if accumulated_length + len(sent_words) > chunk_size and split_idx > 0:
                        break
                    accumulated_length += len(sent_words)
                    split_idx += 1
                
                if split_idx > 0:
                    # Split at sentence boundary
                    first_part = " ".join(proper_sentences[:split_idx])
                    remaining = " ".join(proper_sentences[split_idx:])
                    
                    if first_part.strip():
                        chunks.append(first_part.strip())
                    
                    # Start new chunk with overlap
                    remaining_words = remaining.split()
                    if overlap > 0 and chunks:
                        prev_words = chunks[-1].split()
                        overlap_words = prev_words[-overlap:] if len(prev_words) >= overlap else prev_words
                        current_chunk = overlap_words + remaining_words
                    else:
                        current_chunk = remaining_words
                    current_length = len(current_chunk)
                else:
                    # No good sentence boundary, force split at word boundary
                    words_to_keep = chunk_size - overlap
                    if words_to_keep > 0:
                        first_part = " ".join(current_chunk[:words_to_keep])
                        if first_part.strip():
                            chunks.append(first_part.strip())
                        
                        # Overlap for next chunk
                        if overlap > 0:
                            current_chunk = current_chunk[words_to_keep - overlap:words_to_keep] + current_chunk[words_to_keep:]
                        else:
                            current_chunk = current_chunk[words_to_keep:]
                        current_length = len(current_chunk)
                    else:
                        # Chunk is too small to split, just save it
                        chunk_text = " ".join(current_chunk)
                        if chunk_text.strip():
                            chunks.append(chunk_text.strip())
                        current_chunk = []
                        current_length = 0
                    break
        
        # Add final chunk
        if current_chunk:
            chunk_text = " ".join(current_chunk)
            if chunk_text.strip():
                chunks.append(chunk_text.strip())
        
        # Filter out very short chunks (likely artifacts)
        chunks = [c for c in chunks if len(c.split()) >= 10]  # At least 10 words
        
        return chunks
    
    def _extract_text_from_file(self, file_path: Path) -> str:
        """Extract text from various file formats"""
        suffix = file_path.suffix.lower()
        
        if suffix == '.txt' or suffix == '.md':
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Error reading {file_path.name}: {e}")
                return ""
        
        elif suffix == '.pdf':
            if not PDF_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ PDF support not available for {file_path.name}")
                return ""
            try:
                text = ""
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
                return text.strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ PDF error {file_path.name}: {e}")
                return ""
        
        elif suffix == '.docx':
            if not DOCX_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ DOCX support not available for {file_path.name}")
                return ""
            try:
                doc = docx.Document(file_path)
                text = "\n".join([p.text for p in doc.paragraphs])
                return text.strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ DOCX error {file_path.name}: {e}")
                return ""
        
        elif suffix in ['.xlsx', '.xls']:
            if not EXCEL_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ Excel support not available for {file_path.name}")
                return ""
            try:
                text_parts = []
                # Try using openpyxl first (for .xlsx)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                        for row in sheet.iter_rows(values_only=True):
                            row_text = []
                            for cell in row:
                                if cell is not None:
                                    cell_str = str(cell).strip()
                                    if cell_str:
                                        row_text.append(cell_str)
                            if row_text:
                                text_parts.append(" | ".join(row_text))
                            text_parts.append("\n")
                    return "\n".join(text_parts).strip()
                except Exception as e1:
                    # Fallback to pandas if openpyxl fails
                    try:
                        import pandas as pd
                        excel_file = pd.ExcelFile(file_path)
                        for sheet_name in excel_file.sheet_names:
                            df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                            text_parts.append(df.to_string(index=False))
                            text_parts.append("\n")
                        return "\n".join(text_parts).strip()
                    except Exception as e2:
                        print(f"[Auto-Ingest] ❌ Excel error {file_path.name}: {e1}, {e2}")
                        return ""
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Excel error {file_path.name}: {e}")
                return ""
        
        else:
            print(f"[Auto-Ingest] ⚠️ Unsupported format: {suffix}")
            return ""
    
    def _process_file(self, file_path: Path, force: bool = False) -> bool:
        """Process a single file - extracts text directly from input files
        
        Args:
            file_path: Path to file to process
            force: If True, process even if already in state (for missing files)
        """
        try:
            original_name = file_path.name
            
            # Check if file exists in embeddings metadata
            file_in_embeddings = False
            if self.metadata:
                for meta in self.metadata:
                    if isinstance(meta, dict):
                        meta_file_path = meta.get("file_path", "")
                        if meta_file_path and Path(meta_file_path).name == original_name:
                            file_in_embeddings = True
                            break
                        doc_name = meta.get("document_name", "")
                        if doc_name == original_name or doc_name == file_path.stem:
                            file_in_embeddings = True
                            break
            
            # If file is missing from embeddings, force processing
            if not file_in_embeddings and original_name in self.state.get("processed_files", {}):
                print(f"[Auto-Ingest] 🔄 Re-processing {file_path.name} (missing from embeddings, forcing reprocess)")
                force = True
            
            # Check if file was already processed (unless forced)
            if not force:
                file_hash = self._get_file_hash(file_path)
                if original_name in self.state.get("processed_files", {}):
                    if self.state["processed_files"][original_name]["hash"] == file_hash:
                        if file_in_embeddings:
                            print(f"[Auto-Ingest] ⏭️ Skipping {file_path.name} (already processed, no changes)")
                            return False  # Already processed, no changes
                        else:
                            # In state but not in embeddings - force reprocess
                            print(f"[Auto-Ingest] 🔄 Re-processing {file_path.name} (in state but missing from embeddings)")
                            force = True
                    else:
                        print(f"[Auto-Ingest] 🔄 Re-processing {file_path.name} (file modified)")
            
            print(f"[Auto-Ingest] 📄 Processing: {file_path.name}")
            
            # Get file hash for state tracking
            file_hash = self._get_file_hash(file_path)
            
            # Extract text directly from file
            content = self._extract_text_from_file(file_path)
            
            if not content.strip():
                print(f"[Auto-Ingest] ⚠️ Empty content extracted from {file_path.name}")
                return False
            
            # Extract document name from filename
            doc_name = file_path.stem
            
            # Chunk the content
            chunks = self.chunk_text(content)
            
            if not chunks:
                print(f"[Auto-Ingest] ⚠️ No chunks created from {file_path.name}")
                return False
            
            # Create metadata for each chunk
            chunk_metadata = []
            for i, chunk in enumerate(chunks):
                chunk_metadata.append({
                    "chunk_id": f"{doc_name}_{i}",
                    "document_name": doc_name,
                    "chunk_index": i,
                    "text": chunk,
                    "file_path": str(file_path)
                })
            
            # Add to existing data
            if isinstance(self.chunks, list):
                self.chunks.extend(chunks)
            else:
                # Convert numpy array to list if needed
                self.chunks = list(self.chunks) + chunks
            
            self.metadata.extend(chunk_metadata)
            
            # Update state
            self.state["processed_files"][original_name] = {
                "hash": file_hash,
                "processed_at": time.time(),
                "chunks": len(chunks)
            }
            
            print(f"[Auto-Ingest] ✅ Processed {file_path.name}: {len(chunks)} chunks (total chunks now: {len(self.chunks)})")
            return True
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Error processing {file_path.name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _generate_embeddings(self) -> np.ndarray:
        """Generate embeddings for all chunks"""
        if not self.chunks:
            return np.array([])
        
        print(f"[Auto-Ingest] 🔧 Generating embeddings for {len(self.chunks)} chunks...")
        embeddings = self.model.encode(self.chunks)
        print(f"[Auto-Ingest] ✅ Generated embeddings: {embeddings.shape}")
        return embeddings
    
    def _save_embeddings_cpu_format(self):
        """Save embeddings in CPU FAISS format"""
        if not self.chunks or not self.metadata:
            print("[Auto-Ingest] ⚠️ No data to save")
            return
        
        try:
            # Generate embeddings
            embeddings = self._generate_embeddings()
            
            if embeddings.size == 0:
                print("[Auto-Ingest] ⚠️ No embeddings to save")
                return
            
            # Normalize embeddings for cosine similarity (required for IndexFlatIP)
            embeddings = embeddings.astype('float32')
            faiss.normalize_L2(embeddings)
            print(f"[Auto-Ingest] ✅ Normalized embeddings for cosine similarity")
            
            # Create FAISS index
            index = faiss.IndexFlatIP(self.embedding_dimension)  # Inner product for cosine similarity
            index.add(embeddings)
            
            # Save CPU FAISS format
            faiss.write_index(index, str(self.cpu_embeddings_dir / "faiss_index.bin"))
            
            # Save metadata
            metadata = {
                "total_chunks": len(self.chunks),
                "embedding_dimension": self.embedding_dimension,
                "model_name": self.model_name,
                "chunks": self.chunks,
                "metadata": self.metadata
            }
            
            with open(self.cpu_embeddings_dir / "metadata.pkl", 'wb') as f:
                pickle.dump(metadata, f)
            
            print(f"[Auto-Ingest] ✅ Saved CPU FAISS embeddings: {len(self.chunks)} chunks to {self.cpu_embeddings_dir}")
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to save CPU embeddings: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def scan_and_process(self) -> Dict[str, Any]:
        """Scan input directory and process new/modified files"""
        print(f"[Auto-Ingest] 🔍 Scanning {self.input_dir} for new/modified files...")
        
        # Ensure input directory exists
        if not self.input_dir.exists():
            print(f"[Auto-Ingest] ⚠️ Input directory does not exist: {self.input_dir}")
            print(f"[Auto-Ingest] 💡 Creating directory...")
            self.input_dir.mkdir(parents=True, exist_ok=True)
            return {"processed": 0, "skipped": 0, "errors": 0, "total_chunks": len(self.chunks)}
        
        # Get list of files currently in embeddings/metadata
        files_in_embeddings = set()
        if self.metadata:
            for meta in self.metadata:
                if isinstance(meta, dict):
                    file_path = meta.get("file_path", "")
                    if file_path:
                        files_in_embeddings.add(Path(file_path).name)
                    else:
                        doc_name = meta.get("document_name", "")
                        if doc_name:
                            # Try to match by document name (might need file extension)
                            files_in_embeddings.add(doc_name)
        
        # Also check state file for processed files
        files_in_state = set(self.state.get("processed_files", {}).keys())
        
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        # Find all files in input directory (PDF, DOCX, TXT, MD, XLSX, XLS, etc.)
        supported_extensions = ['.pdf', '.docx', '.txt', '.md', '.xlsx', '.xls']
        input_files = []
        for ext in supported_extensions:
            input_files.extend(self.input_dir.glob(f"*{ext}"))
        
        # Get set of input file names for comparison
        input_file_names = {f.name for f in input_files}
        
        print(f"[Auto-Ingest] 📂 Found {len(input_files)} file(s) in input directory")
        if input_files:
            print(f"[Auto-Ingest] 📋 Files found:")
            for file_path in input_files:
                print(f"[Auto-Ingest]   - {file_path.name}")
        
        # Clean up orphaned embeddings (files in embeddings but missing from input)
        orphaned_files = files_in_embeddings - input_file_names
        if orphaned_files:
            print(f"[Auto-Ingest] 🗑️ Found {len(orphaned_files)} orphaned file(s) in embeddings (missing from input):")
            for orphaned_file in orphaned_files:
                print(f"[Auto-Ingest]   - {orphaned_file}")
            
            # Remove chunks and metadata for orphaned files
            indices_to_remove = []
            for i, meta in enumerate(self.metadata):
                if isinstance(meta, dict):
                    meta_file_path = meta.get("file_path", "")
                    doc_name = meta.get("document_name", "")
                    file_name = None
                    
                    if meta_file_path:
                        file_name = Path(meta_file_path).name
                    elif doc_name:
                        # Check if doc_name matches any orphaned file (with or without extension)
                        for orphaned in orphaned_files:
                            if (doc_name == orphaned or 
                                doc_name == Path(orphaned).stem or
                                doc_name.lower() == orphaned.lower() or
                                doc_name.lower() == Path(orphaned).stem.lower()):
                                file_name = orphaned
                                break
                    
                    if file_name and file_name in orphaned_files:
                        indices_to_remove.append(i)
            
            # Remove in reverse order to maintain indices
            removed_count = 0
            for i in reversed(indices_to_remove):
                if i < len(self.chunks) and i < len(self.metadata):
                    del self.chunks[i]
                    del self.metadata[i]
                    removed_count += 1
            
            if removed_count > 0:
                print(f"[Auto-Ingest] 🗑️ Removed {removed_count} orphaned chunk(s) from {len(orphaned_files)} file(s)")
                # Remove from state
                for orphaned_file in orphaned_files:
                    if orphaned_file in self.state.get("processed_files", {}):
                        del self.state["processed_files"][orphaned_file]
                        print(f"[Auto-Ingest] 🗑️ Removed {orphaned_file} from processed state")
                
                # Update files_in_embeddings set after cleanup
                files_in_embeddings = files_in_embeddings - orphaned_files
                
                # Save cleaned embeddings
                print(f"[Auto-Ingest] 💾 Saving cleaned embeddings...")
                self._save_embeddings_cpu_format()
                self._save_state()
                print(f"[Auto-Ingest] ✅ Cleaned embeddings saved: {len(self.chunks)} chunks remaining")
        
        # Check for missing files (in input but not in embeddings)
        # A file is "missing" if it's in input but not in embeddings (regardless of state)
        # This handles cases where embeddings were deleted or state is out of sync
        missing_files = []
        for file_path in input_files:
            file_name = file_path.name
            if file_name not in files_in_embeddings:
                missing_files.append(file_path)
        
        if missing_files:
            print(f"[Auto-Ingest] ⚠️ Found {len(missing_files)} file(s) in input but missing from embeddings:")
            for file_path in missing_files:
                in_state = file_path.name in files_in_state
                status = "in state but missing from embeddings" if in_state else "not in state or embeddings"
                print(f"[Auto-Ingest]   - {file_path.name} ({status}) - will be processed")
        
        # Process all input files (will skip if already processed and unchanged)
        for file_path in input_files:
            try:
                if self._process_file(file_path):
                    processed_count += 1
                else:
                    skipped_count += 1
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Error processing {file_path.name}: {e}")
                error_count += 1
        
        # Save embeddings if any files were processed
        if processed_count > 0:
            self._save_embeddings_cpu_format()
            self._save_state()
        
        result = {
            "processed": processed_count,
            "skipped": skipped_count,
            "errors": error_count,
            "total_chunks": len(self.chunks)
        }
        
        print(f"[Auto-Ingest] 📊 Scan complete: {processed_count} processed, {skipped_count} skipped, {error_count} errors")
        return result
    
    def load_existing_embeddings(self) -> bool:
        """Load existing embeddings from disk"""
        try:
            index_path = self.cpu_embeddings_dir / "faiss_index.bin"
            metadata_path = self.cpu_embeddings_dir / "metadata.pkl"
            
            if not index_path.exists() or not metadata_path.exists():
                print("[Auto-Ingest] ⚠️ No existing embeddings found")
                return False
            
            # Load metadata
            with open(metadata_path, 'rb') as f:
                metadata = pickle.load(f)
            
            self.chunks = metadata.get("chunks", [])
            self.metadata = metadata.get("metadata", [])
            
            # Extract unique file names from metadata
            unique_files = set()
            for meta in self.metadata:
                if isinstance(meta, dict):
                    file_path = meta.get("file_path", "")
                    if file_path:
                        unique_files.add(Path(file_path).name)
                    else:
                        doc_name = meta.get("document_name", "")
                        if doc_name:
                            unique_files.add(doc_name)
            
            print(f"[Auto-Ingest] ✅ Loaded existing embeddings: {len(self.chunks)} chunks from {len(unique_files)} file(s)")
            if unique_files:
                print(f"[Auto-Ingest] 📚 Files in index:")
                for file_name in sorted(unique_files):
                    # Count chunks per file
                    file_chunks = sum(1 for meta in self.metadata if isinstance(meta, dict) and 
                                    (meta.get("file_path", "").endswith(file_name) or meta.get("document_name") == file_name))
                    print(f"[Auto-Ingest]   - {file_name} ({file_chunks} chunks)")
            return True
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to load existing embeddings: {e}")
            return False
    
    def start_watching(self):
        """Start file system watching"""
        if self.watching:
            print(f"[Auto-Ingest] ⚠️ Already watching: {self.input_dir}")
            return
        
        print(f"[Auto-Ingest] 👀 Starting file watcher for: {self.input_dir}")
        try:
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler
            
            class GuidelineHandler(FileSystemEventHandler):
                def __init__(self, auto_ingest):
                    self.auto_ingest = auto_ingest
                    # Supported file extensions for auto-ingestion
                    self.supported_extensions = ['.pdf', '.docx', '.txt', '.md', '.xlsx', '.xls']
                
                def _is_supported_file(self, path):
                    """Check if file has a supported extension"""
                    return any(path.lower().endswith(ext) for ext in self.supported_extensions)
                
                def on_created(self, event):
                    if not event.is_directory and self._is_supported_file(event.src_path):
                        file_name = Path(event.src_path).name
                        print(f"[Auto-Ingest] 📥 New file detected: {file_name}")
                        time.sleep(1)  # Wait for file to be fully written
                        print(f"[Auto-Ingest] 🔄 Processing new file: {file_name}")
                        self.auto_ingest.scan_and_process()
                
                def on_modified(self, event):
                    if not event.is_directory and self._is_supported_file(event.src_path):
                        file_name = Path(event.src_path).name
                        print(f"[Auto-Ingest] 📝 File modified: {file_name}")
                        time.sleep(1)  # Wait for file to be fully written
                        print(f"[Auto-Ingest] 🔄 Re-processing modified file: {file_name}")
                        self.auto_ingest.scan_and_process()
            
            self.observer = Observer()
            self.observer.schedule(GuidelineHandler(self), str(self.input_dir), recursive=False)
            self.observer.start()
            self.watching = True
            
            print(f"[Auto-Ingest] ✅ File watcher started for: {self.input_dir}")
            print(f"[Auto-Ingest] 👁️ Watching for changes to: PDF, DOCX, TXT, MD, XLSX, XLS files")
            
        except ImportError:
            print("[Auto-Ingest] ⚠️ Watchdog not available, file watching disabled")
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to start watching: {e}")
    
    def stop_watching(self):
        """Stop file system watching"""
        if self.observer and self.watching:
            self.observer.stop()
            self.observer.join()
            self.watching = False
            print("[Auto-Ingest] 🛑 Stopped file watching")

if __name__ == "__main__":
    # Test the auto-ingestion system
    auto_ingest = CPUFAISSAutoIngest()
    
    # Load existing embeddings
    auto_ingest.load_existing_embeddings()
    
    # Scan and process
    result = auto_ingest.scan_and_process()
    print(f"Result: {result}")
