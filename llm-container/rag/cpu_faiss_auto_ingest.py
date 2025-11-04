#!/usr/bin/env python3
"""
CPU FAISS Auto-Ingestion System
Monitors data/input/ for new guideline files and automatically updates embeddings
"""

import os
import time
import json
import pickle
import hashlib
from pathlib import Path
from typing import Dict, List, Any
import numpy as np
import faiss
from sentence_transformers import SentenceTransformer
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    try:
        import pandas as pd
        EXCEL_SUPPORT = True
    except ImportError:
        EXCEL_SUPPORT = False
        print("[Auto-Ingest] ⚠️ Excel support not available. Install openpyxl or pandas")

# PDF and DOCX support
try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("[Auto-Ingest] ⚠️ PDF support not available. Install PyPDF2")

try:
    import docx
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False
    print("[Auto-Ingest] ⚠️ DOCX support not available. Install python-docx")

class CPUFAISSAutoIngest:
    """Auto-ingestion system for CPU FAISS"""
    
    def __init__(self):
        self.input_dir = Path("../data/input")  # Relative to container
        self.parsed_dir = Path("../data/parsed")  # Read parsed text from GPU RAG extraction
        self.cpu_embeddings_dir = Path("../data/embeddings")  # Relative to container
        self.model_name = "all-distilroberta-v1"
        self.embedding_dimension = 768
        
        # Initialize components
        self.model = None
        self.index = None
        self.chunks = []
        self.metadata = []
        self.state_file = Path("data/cpu_ingest_state.json")
        self.state = self._load_state()
        
        # Ensure directories exist
        self.input_dir.mkdir(parents=True, exist_ok=True)
        self.parsed_dir.mkdir(parents=True, exist_ok=True)
        self.cpu_embeddings_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize model
        self._load_model()
        
        # Initialize file watcher
        self.observer = None
        self.watching = False
    
    def _load_model(self):
        """Load the sentence transformer model"""
        try:
            self.model = SentenceTransformer(self.model_name)
            print(f"[Auto-Ingest] ✅ Loaded model: {self.model_name}")
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to load model: {e}")
            raise
    
    def _load_state(self) -> Dict:
        """Load ingestion state"""
        if self.state_file.exists():
            try:
                with open(self.state_file, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"[Auto-Ingest] ⚠️ Failed to load state: {e}")
        return {"processed_files": {}, "last_scan": None}
    
    def _save_state(self):
        """Save ingestion state"""
        try:
            with open(self.state_file, 'w') as f:
                json.dump(self.state, f, indent=2)
        except Exception as e:
            print(f"[Auto-Ingest] ⚠️ Failed to save state: {e}")
    
    def _get_file_hash(self, file_path: Path) -> str:
        """Get file hash for change detection"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception as e:
            print(f"[Auto-Ingest] ⚠️ Failed to get hash for {file_path}: {e}")
            return ""
    
    def chunk_text(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """Chunk text into overlapping segments"""
        words = text.split()
        chunks = []
        
        for i in range(0, len(words), chunk_size - overlap):
            chunk = " ".join(words[i:i + chunk_size])
            if chunk.strip():
                chunks.append(chunk.strip())
        
        return chunks
    
    def _extract_text_from_file(self, file_path: Path) -> str:
        """Extract text from various file formats"""
        suffix = file_path.suffix.lower()
        
        if suffix == '.txt' or suffix == '.md':
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Error reading {file_path.name}: {e}")
                return ""
        
        elif suffix == '.pdf':
            if not PDF_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ PDF support not available for {file_path.name}")
                return ""
            try:
                text = ""
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
                return text.strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ PDF error {file_path.name}: {e}")
                return ""
        
        elif suffix == '.docx':
            if not DOCX_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ DOCX support not available for {file_path.name}")
                return ""
            try:
                doc = docx.Document(file_path)
                text = "\n".join([p.text for p in doc.paragraphs])
                return text.strip()
            except Exception as e:
                print(f"[Auto-Ingest] ❌ DOCX error {file_path.name}: {e}")
                return ""
        
        elif suffix in ['.xlsx', '.xls']:
            if not EXCEL_SUPPORT:
                print(f"[Auto-Ingest] ⚠️ Excel support not available for {file_path.name}")
                return ""
            try:
                text_parts = []
                # Try using openpyxl first (for .xlsx)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                        for row in sheet.iter_rows(values_only=True):
                            row_text = []
                            for cell in row:
                                if cell is not None:
                                    cell_str = str(cell).strip()
                                    if cell_str:
                                        row_text.append(cell_str)
                            if row_text:
                                text_parts.append(" | ".join(row_text))
                            text_parts.append("\n")
                    return "\n".join(text_parts).strip()
                except Exception as e1:
                    # Fallback to pandas if openpyxl fails
                    try:
                        import pandas as pd
                        excel_file = pd.ExcelFile(file_path)
                        for sheet_name in excel_file.sheet_names:
                            df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                            text_parts.append(df.to_string(index=False))
                            text_parts.append("\n")
                        return "\n".join(text_parts).strip()
                    except Exception as e2:
                        print(f"[Auto-Ingest] ❌ Excel error {file_path.name}: {e1}, {e2}")
                        return ""
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Excel error {file_path.name}: {e}")
                return ""
        
        else:
            print(f"[Auto-Ingest] ⚠️ Unsupported format: {suffix}")
            return ""
    
    def _process_file(self, file_path: Path) -> bool:
        """Process a single file - extracts text directly from input files"""
        try:
            # Check if file was already processed
            file_hash = self._get_file_hash(file_path)
            original_name = file_path.name
            if original_name in self.state["processed_files"]:
                if self.state["processed_files"][original_name]["hash"] == file_hash:
                    return False  # Already processed, no changes
            
            print(f"[Auto-Ingest] 📄 Processing: {file_path.name}")
            
            # Extract text directly from file
            content = self._extract_text_from_file(file_path)
            
            if not content.strip():
                print(f"[Auto-Ingest] ⚠️ Empty content extracted from {file_path.name}")
                return False
            
            # Extract document name from filename
            doc_name = file_path.stem
            
            # Chunk the content
            chunks = self.chunk_text(content)
            
            if not chunks:
                print(f"[Auto-Ingest] ⚠️ No chunks created from {file_path.name}")
                return False
            
            # Create metadata for each chunk
            chunk_metadata = []
            for i, chunk in enumerate(chunks):
                chunk_metadata.append({
                    "chunk_id": f"{doc_name}_{i}",
                    "document_name": doc_name,
                    "chunk_index": i,
                    "text": chunk,
                    "file_path": str(file_path),
                    "parsed_file": str(parsed_file)
                })
            
            # Add to existing data
            if isinstance(self.chunks, list):
                self.chunks.extend(chunks)
            else:
                # Convert numpy array to list if needed
                self.chunks = list(self.chunks) + chunks
            
            self.metadata.extend(chunk_metadata)
            
            # Update state
            self.state["processed_files"][original_name] = {
                "hash": file_hash,
                "processed_at": time.time(),
                "chunks": len(chunks)
            }
            
            print(f"[Auto-Ingest] ✅ Processed {file_path.name}: {len(chunks)} chunks")
            return True
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Error processing {file_path.name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _generate_embeddings(self) -> np.ndarray:
        """Generate embeddings for all chunks"""
        if not self.chunks:
            return np.array([])
        
        print(f"[Auto-Ingest] 🔧 Generating embeddings for {len(self.chunks)} chunks...")
        embeddings = self.model.encode(self.chunks)
        print(f"[Auto-Ingest] ✅ Generated embeddings: {embeddings.shape}")
        return embeddings
    
    def _save_embeddings_cpu_format(self):
        """Save embeddings in CPU FAISS format"""
        if not self.chunks or not self.metadata:
            print("[Auto-Ingest] ⚠️ No data to save")
            return
        
        try:
            # Generate embeddings
            embeddings = self._generate_embeddings()
            
            if embeddings.size == 0:
                print("[Auto-Ingest] ⚠️ No embeddings to save")
                return
            
            # Create FAISS index
            index = faiss.IndexFlatIP(self.embedding_dimension)  # Inner product for cosine similarity
            index.add(embeddings.astype('float32'))
            
            # Save CPU FAISS format
            faiss.write_index(index, str(self.cpu_embeddings_dir / "faiss_index.bin"))
            
            # Save metadata
            metadata = {
                "total_chunks": len(self.chunks),
                "embedding_dimension": self.embedding_dimension,
                "model_name": self.model_name,
                "chunks": self.chunks,
                "metadata": self.metadata
            }
            
            with open(self.cpu_embeddings_dir / "metadata.pkl", 'wb') as f:
                pickle.dump(metadata, f)
            
            print(f"[Auto-Ingest] ✅ Saved CPU FAISS embeddings: {len(self.chunks)} chunks")
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to save CPU embeddings: {e}")
            raise
    
    def scan_and_process(self) -> Dict[str, Any]:
        """Scan input directory and process new/modified files from parsed text"""
        print("[Auto-Ingest] 🔍 Scanning for new/modified files...")
        
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        # Find all files in input directory (PDF, DOCX, TXT, MD, XLSX, XLS, etc.)
        # We'll process them by reading their parsed text counterparts
        supported_extensions = ['.pdf', '.docx', '.txt', '.md', '.xlsx', '.xls']
        input_files = []
        for ext in supported_extensions:
            input_files.extend(self.input_dir.glob(f"*{ext}"))
        
        for file_path in input_files:
            try:
                if self._process_file(file_path):
                    processed_count += 1
                else:
                    skipped_count += 1
            except Exception as e:
                print(f"[Auto-Ingest] ❌ Error processing {file_path.name}: {e}")
                error_count += 1
        
        # Save embeddings if any files were processed
        if processed_count > 0:
            self._save_embeddings_cpu_format()
            self._save_state()
        
        result = {
            "processed": processed_count,
            "skipped": skipped_count,
            "errors": error_count,
            "total_chunks": len(self.chunks)
        }
        
        print(f"[Auto-Ingest] 📊 Scan complete: {processed_count} processed, {skipped_count} skipped, {error_count} errors")
        return result
    
    def load_existing_embeddings(self) -> bool:
        """Load existing embeddings from disk"""
        try:
            index_path = self.cpu_embeddings_dir / "faiss_index.bin"
            metadata_path = self.cpu_embeddings_dir / "metadata.pkl"
            
            if not index_path.exists() or not metadata_path.exists():
                print("[Auto-Ingest] ⚠️ No existing embeddings found")
                return False
            
            # Load metadata
            with open(metadata_path, 'rb') as f:
                metadata = pickle.load(f)
            
            self.chunks = metadata.get("chunks", [])
            self.metadata = metadata.get("metadata", [])
            
            print(f"[Auto-Ingest] ✅ Loaded existing embeddings: {len(self.chunks)} chunks")
            return True
            
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to load existing embeddings: {e}")
            return False
    
    def start_watching(self):
        """Start file system watching"""
        if self.watching:
            return
        
        try:
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler
            
            class GuidelineHandler(FileSystemEventHandler):
                def __init__(self, auto_ingest):
                    self.auto_ingest = auto_ingest
                
                def on_created(self, event):
                    if not event.is_directory and event.src_path.endswith('.txt'):
                        self.auto_ingest.scan_and_process()
                
                def on_modified(self, event):
                    if not event.is_directory and event.src_path.endswith('.txt'):
                        self.auto_ingest.scan_and_process()
            
            self.observer = Observer()
            self.observer.schedule(GuidelineHandler(self), str(self.input_dir), recursive=False)
            self.observer.start()
            self.watching = True
            
            print("[Auto-Ingest] 👀 Started file watching")
            
        except ImportError:
            print("[Auto-Ingest] ⚠️ Watchdog not available, file watching disabled")
        except Exception as e:
            print(f"[Auto-Ingest] ❌ Failed to start watching: {e}")
    
    def stop_watching(self):
        """Stop file system watching"""
        if self.observer and self.watching:
            self.observer.stop()
            self.observer.join()
            self.watching = False
            print("[Auto-Ingest] 🛑 Stopped file watching")

if __name__ == "__main__":
    # Test the auto-ingestion system
    auto_ingest = CPUFAISSAutoIngest()
    
    # Load existing embeddings
    auto_ingest.load_existing_embeddings()
    
    # Scan and process
    result = auto_ingest.scan_and_process()
    print(f"Result: {result}")
