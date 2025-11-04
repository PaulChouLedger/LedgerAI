#!/usr/bin/env python3
"""
Auto-ingest functionality for RAG container
Processes files from data/input and updates embeddings
"""
import os
import json
import hashlib
import shutil
import numpy as np
import faiss
from pathlib import Path
from typing import List, Dict
import PyPDF2
import docx
import re
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    try:
        import pandas as pd
        EXCEL_SUPPORT = True
    except ImportError:
        EXCEL_SUPPORT = False
        print("[Ingest] ⚠️ Excel support not available. Install openpyxl or pandas: pip install openpyxl")

class AutoIngest:
    def __init__(self, rag_instance):
        """
        Initialize auto-ingest using existing RAG instance
        
        Args:
            rag_instance: Existing RAG object (reuses encoder and index)
        """
        self.rag = rag_instance
        self.input_dir = Path("data/input")
        self.parsed_dir = Path("data/parsed")
        self.embeddings_dir = Path("data/embeddings")
        self.state_file = Path("data/ingest_state.json")
        
        # Chunking parameters
        self.chunk_size = 400
        self.chunk_overlap = 50
        
        # Create directories
        self.input_dir.mkdir(parents=True, exist_ok=True)
        self.parsed_dir.mkdir(parents=True, exist_ok=True)
        
        # Load state
        self.state = self.load_state()
        
        print("[Ingest] ✅ Auto-ingest initialized")
    
    def load_state(self) -> Dict:
        """Load processing state"""
        if self.state_file.exists():
            try:
                with open(self.state_file, 'r') as f:
                    return json.load(f)
            except:
                pass
        return {"processed_files": {}, "total_chunks": 0}
    
    def save_state(self):
        """Save processing state"""
        try:
            with open(self.state_file, 'w') as f:
                json.dump(self.state, f, indent=2)
        except Exception as e:
            print(f"[Ingest] ❌ Error saving state: {e}")
    
    def get_file_hash(self, file_path: Path) -> str:
        """Calculate MD5 hash"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF"""
        try:
            text = ""
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            print(f"[Ingest] ❌ PDF error {pdf_path.name}: {e}")
            return ""
    
    def extract_text_from_docx(self, docx_path: Path) -> str:
        """Extract text from DOCX"""
        try:
            doc = docx.Document(docx_path)
            text = "\n".join([p.text for p in doc.paragraphs])
            return text.strip()
        except Exception as e:
            print(f"[Ingest] ❌ DOCX error {docx_path.name}: {e}")
            return ""
    
    def extract_text_from_txt(self, txt_path: Path) -> str:
        """Extract text from TXT/MD"""
        try:
            with open(txt_path, 'r', encoding='utf-8') as f:
                return f.read().strip()
        except Exception as e:
            print(f"[Ingest] ❌ TXT error {txt_path.name}: {e}")
            return ""
    
    def extract_text_from_excel(self, excel_path: Path) -> str:
        """Extract text from Excel files (.xlsx, .xls)"""
        if not EXCEL_SUPPORT:
            print(f"[Ingest] ❌ Excel support not available. Install openpyxl: pip install openpyxl")
            return ""
        
        try:
            text_parts = []
            
            # Try using openpyxl first (for .xlsx)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = []
                        for cell in row:
                            if cell is not None:
                                # Convert cell value to string, handling different types
                                cell_str = str(cell).strip()
                                if cell_str:
                                    row_text.append(cell_str)
                        
                        if row_text:
                            text_parts.append(" | ".join(row_text))
                        text_parts.append("\n")
                
                return "\n".join(text_parts).strip()
            except Exception as e1:
                # Fallback to pandas if openpyxl fails
                try:
                    import pandas as pd
                    # Read all sheets
                    excel_file = pd.ExcelFile(excel_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                        # Convert DataFrame to text representation
                        text_parts.append(df.to_string(index=False))
                        text_parts.append("\n")
                    
                    return "\n".join(text_parts).strip()
                except Exception as e2:
                    print(f"[Ingest] ❌ Excel error {excel_path.name}: {e1}, {e2}")
                    return ""
                    
        except Exception as e:
            print(f"[Ingest] ❌ Excel error {excel_path.name}: {e}")
            return ""
    
    def extract_text(self, file_path: Path) -> str:
        """Extract text based on file extension"""
        suffix = file_path.suffix.lower()
        if suffix == '.pdf':
            return self.extract_text_from_pdf(file_path)
        elif suffix == '.docx':
            return self.extract_text_from_docx(file_path)
        elif suffix in ['.txt', '.md']:
            return self.extract_text_from_txt(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return self.extract_text_from_excel(file_path)
        else:
            print(f"[Ingest] ⚠️ Unsupported format: {suffix}")
            return ""
    
    def chunk_text(self, text: str) -> List[str]:
        """Split text into overlapping chunks"""
        if len(text) <= self.chunk_size:
            return [text.strip()]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            # Try to break at sentence boundary
            if end < len(text):
                search_start = max(start + self.chunk_size - 100, start)
                sentence_end = -1
                
                for i in range(end, search_start, -1):
                    if text[i] in '.!?':
                        if i + 1 < len(text) and text[i + 1] in ' \n\t':
                            sentence_end = i + 1
                            break
                
                if sentence_end > 0:
                    end = sentence_end
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - self.chunk_overlap
            if start >= len(text):
                break
        
        return chunks
    
    def process_file(self, file_path: Path) -> bool:
        """Process a single file and update RAG"""
        try:
            print(f"\n[Ingest] 📄 Processing: {file_path.name}")
            
            # Check if already processed
            file_hash = self.get_file_hash(file_path)
            if file_path.name in self.state["processed_files"]:
                if self.state["processed_files"][file_path.name]["hash"] == file_hash:
                    print(f"[Ingest] ⏭️ Already processed (unchanged): {file_path.name}")
                    return False
            
            # Extract text
            text = self.extract_text(file_path)
            if not text:
                print(f"[Ingest] ⚠️ No text extracted from {file_path.name}")
                return False
            
            # Save parsed text
            parsed_path = self.parsed_dir / f"{file_path.stem}.txt"
            with open(parsed_path, 'w', encoding='utf-8') as f:
                f.write(text)
            print(f"[Ingest] ✅ Saved parsed text: {parsed_path.name}")
            
            # File successfully parsed - text is saved to data/parsed/
            # The HOST will handle embedding generation (has working FAISS)
            print(f"[Ingest] ✅ Parsed text saved - ready for embedding generation by host")
            
            # Update state
            self.state["processed_files"][file_path.name] = {
                "hash": file_hash,
                "timestamp": str(Path(file_path).stat().st_mtime)
            }
            self.save_state()
            
            print(f"[Ingest] ✅ Processed {file_path.name}")
            
            return True
            
        except Exception as e:
            print(f"[Ingest] ❌ Error processing {file_path.name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def scan_and_process(self) -> Dict:
        """Scan input directory and process new/changed files"""
        if not self.input_dir.exists():
            return {"processed": 0, "skipped": 0, "errors": 0}
        
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        # Find all supported files
        supported_extensions = ['.pdf', '.docx', '.txt', '.md']
        files = [f for f in self.input_dir.iterdir() 
                if f.is_file() and f.suffix.lower() in supported_extensions]
        
        if not files:
            print("[Ingest] ℹ️ No files found in input directory")
            return {"processed": 0, "skipped": 0, "errors": 0}
        
        print(f"\n[Ingest] 🔍 Found {len(files)} file(s) to check")
        
        for file_path in files:
            try:
                if self.process_file(file_path):
                    processed_count += 1
                else:
                    skipped_count += 1
            except Exception as e:
                print(f"[Ingest] ❌ Error with {file_path.name}: {e}")
                error_count += 1
        
        result = {
            "processed": processed_count,
            "skipped": skipped_count,
            "errors": error_count,
            "total_chunks": self.rag.index.ntotal if self.rag.index else 0
        }
        
        if processed_count > 0:
            print(f"\n[Ingest] ✅ Scan complete: {processed_count} file(s) extracted to data/parsed/")
            print(f"[Ingest] 💡 Host will build embeddings from {processed_count} new file(s)")
        
        return result

