#!/usr/bin/env python3
"""
CPU FAISS Auto-Ingestion for llm-medical-container

This implements automatic ingestion similar to the GPU FAISS, but for CPU FAISS.
Monitors data/input/ for new files and automatically processes them.
"""

import os
import json
import pickle
import numpy as np
import faiss
import time
import hashlib
from pathlib import Path
from typing import List, Dict, Any
from sentence_transformers import SentenceTransformer
import threading
import logging

# File extraction dependencies
try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    logger.warning("PDF support not available. Install PyPDF2: pip install PyPDF2")

try:
    import docx
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False
    logger.warning("DOCX support not available. Install python-docx: pip install python-docx")

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    try:
        import pandas as pd
        EXCEL_SUPPORT = True
    except ImportError:
        EXCEL_SUPPORT = False
        logger.warning("Excel support not available. Install openpyxl: pip install openpyxl")

logger = logging.getLogger(__name__)

class CPUFAISSAutoIngest:
    """
    Auto-ingestion system for CPU FAISS in llm-medical-container
    Similar to GPU FAISS auto-ingestion but optimized for CPU operations
    """
    
    def __init__(self, 
                 input_dir: str = "data/input",
                 embeddings_dir: str = "data/embeddings", 
                 cpu_embeddings_dir: str = "llm-medical-container/data/embeddings",
                 model_name: str = "all-distilroberta-v1",
                 chunk_size: int = 400,
                 chunk_overlap: int = 50):
        
        self.input_dir = Path(input_dir)
        self.embeddings_dir = Path(embeddings_dir)
        self.cpu_embeddings_dir = Path(cpu_embeddings_dir)
        self.model_name = model_name
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        
        # Create directories
        self.embeddings_dir.mkdir(parents=True, exist_ok=True)
        self.cpu_embeddings_dir.mkdir(parents=True, exist_ok=True)
        
        # State tracking
        self.state_file = Path("data/cpu_ingest_state.json")
        self.state = self.load_state()
        
        # Initialize components
        self.encoder = None
        self.index = None
        self.chunks = []
        self.chunk_metadata = []
        
        # File watching
        self.watching = False
        self.watch_thread = None
        
        logger.info(f"[CPU Auto-Ingest] 📂 Input: {self.input_dir}")
        logger.info(f"[CPU Auto-Ingest] 📂 GPU embeddings: {self.embeddings_dir}")
        logger.info(f"[CPU Auto-Ingest] 📂 CPU embeddings: {self.cpu_embeddings_dir}")
    
    def load_state(self) -> Dict:
        """Load processing state"""
        if self.state_file.exists():
            try:
                with open(self.state_file, 'r') as f:
                    return json.load(f)
            except:
                pass
        return {"processed_files": {}, "total_chunks": 0}
    
    def save_state(self):
        """Save processing state"""
        try:
            with open(self.state_file, 'w') as f:
                json.dump(self.state, f, indent=2)
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Error saving state: {e}")
    
    def get_file_hash(self, file_path: Path) -> str:
        """Calculate MD5 hash of file"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except:
            return ""
    
    def initialize_encoder(self):
        """Initialize sentence transformer model"""
        if self.encoder is None:
            logger.info(f"[CPU Auto-Ingest] 🔧 Loading encoder: {self.model_name}")
            self.encoder = SentenceTransformer(self.model_name)
            logger.info(f"[CPU Auto-Ingest] ✅ Encoder loaded")
    
    def load_existing_embeddings(self):
        """Load existing embeddings if available"""
        try:
            # Try to load GPU format first
            gpu_index_path = self.embeddings_dir / "index.faiss"
            gpu_chunks_path = self.embeddings_dir / "doc_chunks.npy"
            
            if gpu_index_path.exists() and gpu_chunks_path.exists():
                logger.info(f"[CPU Auto-Ingest] 🔧 Loading existing GPU embeddings...")
                
                # Load GPU format
                self.index = faiss.read_index(str(gpu_index_path))
                self.chunks = np.load(gpu_chunks_path, allow_pickle=True)
                
                # Load metadata
                metadata_path = self.embeddings_dir / "chunk_metadata.json"
                if metadata_path.exists():
                    with open(metadata_path, 'r') as f:
                        metadata = json.load(f)
                        self.chunk_metadata = metadata
                
                logger.info(f"[CPU Auto-Ingest] ✅ Loaded existing embeddings: {self.index.ntotal} vectors")
                return True
                
        except Exception as e:
            logger.warning(f"[CPU Auto-Ingest] ⚠️ Could not load existing embeddings: {e}")
        
        # Try to load CPU format
        try:
            cpu_index_path = self.cpu_embeddings_dir / "faiss_index.bin"
            cpu_metadata_path = self.cpu_embeddings_dir / "metadata.pkl"
            
            if cpu_index_path.exists() and cpu_metadata_path.exists():
                logger.info(f"[CPU Auto-Ingest] 🔧 Loading existing CPU embeddings...")
                
                self.index = faiss.read_index(str(cpu_index_path))
                
                with open(cpu_metadata_path, 'rb') as f:
                    data = pickle.load(f)
                    self.chunks = data.get('chunks', [])
                    self.chunk_metadata = data.get('metadata', {})
                
                logger.info(f"[CPU Auto-Ingest] ✅ Loaded existing CPU embeddings: {self.index.ntotal} vectors")
                return True
                
        except Exception as e:
            logger.warning(f"[CPU Auto-Ingest] ⚠️ Could not load existing CPU embeddings: {e}")
        
        # No existing embeddings found
        logger.info(f"[CPU Auto-Ingest] ⚠️ No existing embeddings found, will create new ones")
        return False
    
    def chunk_text(self, text: str, condition_name: str) -> List[Dict]:
        """Split text into overlapping chunks"""
        chunks = []
        
        # Split by paragraphs first
        paragraphs = text.split('\n\n')
        current_chunk = ""
        
        for paragraph in paragraphs:
            if len(current_chunk) + len(paragraph) <= self.chunk_size:
                current_chunk += paragraph + "\n\n"
            else:
                if current_chunk.strip():
                    chunks.append({
                        'text': current_chunk.strip(),
                        'condition': condition_name,
                        'chunk_id': len(chunks)
                    })
                current_chunk = paragraph + "\n\n"
        
        # Add final chunk
        if current_chunk.strip():
            chunks.append({
                'text': current_chunk.strip(),
                'condition': condition_name,
                'chunk_id': len(chunks)
            })
        
        return chunks
    
    def process_file(self, file_path: Path) -> bool:
        """Process a single guideline file"""
        try:
            logger.info(f"[CPU Auto-Ingest] 📄 Processing: {file_path.name}")
            
            # Check if already processed
            file_hash = self.get_file_hash(file_path)
            if file_path.name in self.state["processed_files"]:
                if self.state["processed_files"][file_path.name]["hash"] == file_hash:
                    logger.info(f"[CPU Auto-Ingest] ⏭️ Already processed (unchanged): {file_path.name}")
                    return False
            
            # Extract text based on file type
            content = self._extract_text(file_path)
            if not content:
                logger.warning(f"[CPU Auto-Ingest] ⚠️ No content extracted from {file_path.name}")
                return False
            
            # Extract condition name from filename
            condition_name = file_path.stem.replace("GUIDELINE_", "")
            
            # Chunk the text
            new_chunks = self.chunk_text(content, condition_name)
            
            # Add to existing chunks
            if isinstance(self.chunks, list):
                self.chunks.extend(new_chunks)
            else:
                # Convert numpy array to list if needed
                self.chunks = list(self.chunks) + new_chunks
            
            # Update state
            self.state["processed_files"][file_path.name] = {
                "hash": file_hash,
                "timestamp": time.time(),
                "chunks_added": len(new_chunks)
            }
            self.state["total_chunks"] = len(self.chunks)
            
            logger.info(f"[CPU Auto-Ingest] ✅ Processed {file_path.name}: {len(new_chunks)} chunks")
            return True
            
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Error processing {file_path.name}: {e}")
            return False
    
    def _extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF"""
        if not PDF_SUPPORT:
            return ""
        try:
            text = ""
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ PDF error {pdf_path.name}: {e}")
            return ""
    
    def _extract_text_from_docx(self, docx_path: Path) -> str:
        """Extract text from DOCX"""
        if not DOCX_SUPPORT:
            return ""
        try:
            doc = docx.Document(docx_path)
            text = "\n".join([p.text for p in doc.paragraphs])
            return text.strip()
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ DOCX error {docx_path.name}: {e}")
            return ""
    
    def _extract_text_from_excel(self, excel_path: Path) -> str:
        """Extract text from Excel files (.xlsx, .xls)"""
        if not EXCEL_SUPPORT:
            return ""
        try:
            text_parts = []
            # Try using openpyxl first (for .xlsx)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = []
                        for cell in row:
                            if cell is not None:
                                cell_str = str(cell).strip()
                                if cell_str:
                                    row_text.append(cell_str)
                        if row_text:
                            text_parts.append(" | ".join(row_text))
                        text_parts.append("\n")
                return "\n".join(text_parts).strip()
            except Exception as e1:
                # Fallback to pandas if openpyxl fails
                try:
                    import pandas as pd
                    excel_file = pd.ExcelFile(excel_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                        text_parts.append(df.to_string(index=False))
                        text_parts.append("\n")
                    return "\n".join(text_parts).strip()
                except Exception as e2:
                    logger.error(f"[CPU Auto-Ingest] ❌ Excel error {excel_path.name}: {e1}, {e2}")
                    return ""
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Excel error {excel_path.name}: {e}")
            return ""
    
    def _extract_text(self, file_path: Path) -> str:
        """Extract text based on file extension"""
        suffix = file_path.suffix.lower()
        if suffix == '.pdf':
            return self._extract_text_from_pdf(file_path)
        elif suffix == '.docx':
            return self._extract_text_from_docx(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return self._extract_text_from_excel(file_path)
        elif suffix in ['.txt', '.md']:
            # Read plain text files
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                logger.error(f"[CPU Auto-Ingest] ❌ TXT error {file_path.name}: {e}")
                return ""
        else:
            logger.warning(f"[CPU Auto-Ingest] ⚠️ Unsupported format: {suffix}")
            return ""
    
    def rebuild_embeddings(self):
        """Rebuild all embeddings from current chunks"""
        if not self.chunks:
            logger.warning(f"[CPU Auto-Ingest] ⚠️ No chunks to process")
            return False
        
        try:
            logger.info(f"[CPU Auto-Ingest] 🔧 Rebuilding embeddings for {len(self.chunks)} chunks...")
            
            # Initialize encoder if needed
            self.initialize_encoder()
            
            # Create embeddings
            texts = [chunk['text'] for chunk in self.chunks]
            embeddings = self.encoder.encode(texts, convert_to_numpy=True)
            
            # Normalize for cosine similarity
            faiss.normalize_L2(embeddings)
            
            # Create FAISS index
            dimension = embeddings.shape[1]
            self.index = faiss.IndexFlatIP(dimension)
            self.index.add(embeddings)
            
            logger.info(f"[CPU Auto-Ingest] ✅ Created embeddings: {embeddings.shape}")
            
            # Save GPU format
            self.save_gpu_format(embeddings)
            
            # Save CPU format
            self.save_cpu_format(embeddings)
            
            # Save state
            self.save_state()
            
            logger.info(f"[CPU Auto-Ingest] ✅ Embeddings rebuilt and saved")
            return True
            
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Error rebuilding embeddings: {e}")
            return False
    
    def save_gpu_format(self, embeddings: np.ndarray):
        """Save embeddings in GPU FAISS format"""
        try:
            # Save FAISS index
            index_path = self.embeddings_dir / "index.faiss"
            faiss.write_index(self.index, str(index_path))
            
            # Save chunks
            chunks_path = self.embeddings_dir / "doc_chunks.npy"
            np.save(chunks_path, self.chunks)
            
            # Save vectors
            vectors_path = self.embeddings_dir / "vectors.npy"
            np.save(vectors_path, embeddings)
            
            # Save metadata
            metadata = {
                'total_chunks': len(self.chunks),
                'embedding_dimension': embeddings.shape[1],
                'model_name': self.model_name,
                'created_at': time.time(),
                'conditions': list(set(chunk['condition'] for chunk in self.chunks))
            }
            
            metadata_path = self.embeddings_dir / "chunk_metadata.json"
            with open(metadata_path, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            logger.info(f"[CPU Auto-Ingest] ✅ Saved GPU format: {index_path}")
            
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Error saving GPU format: {e}")
    
    def save_cpu_format(self, embeddings: np.ndarray):
        """Save embeddings in CPU FAISS format"""
        try:
            # Save FAISS index
            cpu_index_path = self.cpu_embeddings_dir / "faiss_index.bin"
            faiss.write_index(self.index, str(cpu_index_path))
            
            # Prepare metadata for pickle format
            cpu_metadata = {
                'chunks': self.chunks,
                'metadata': {
                    'total_chunks': len(self.chunks),
                    'embedding_dimension': embeddings.shape[1],
                    'model_name': self.model_name,
                    'created_at': time.time(),
                    'conditions': list(set(chunk['condition'] for chunk in self.chunks))
                },
                'total_chunks': len(self.chunks),
                'embedding_dimension': embeddings.shape[1],
                'model_name': self.model_name
            }
            
            # Save metadata
            cpu_metadata_path = self.cpu_embeddings_dir / "metadata.pkl"
            with open(cpu_metadata_path, 'wb') as f:
                pickle.dump(cpu_metadata, f)
            
            logger.info(f"[CPU Auto-Ingest] ✅ Saved CPU format: {cpu_index_path}")
            
        except Exception as e:
            logger.error(f"[CPU Auto-Ingest] ❌ Error saving CPU format: {e}")
    
    def scan_and_process(self) -> Dict:
        """Scan input directory and process new/changed files"""
        logger.info(f"[CPU Auto-Ingest] 🔍 Scanning {self.input_dir} for new files...")
        
        if not self.input_dir.exists():
            logger.warning(f"[CPU Auto-Ingest] ⚠️ Input directory not found: {self.input_dir}")
            return {"processed": 0, "skipped": 0, "errors": 0}
        
        # Find all guideline files
        guideline_files = list(self.input_dir.glob("GUIDELINE_*.txt"))
        logger.info(f"[CPU Auto-Ingest] 📚 Found {len(guideline_files)} guideline files")
        
        processed = 0
        skipped = 0
        errors = 0
        
        for file_path in guideline_files:
            try:
                if self.process_file(file_path):
                    processed += 1
                else:
                    skipped += 1
            except Exception as e:
                logger.error(f"[CPU Auto-Ingest] ❌ Error with {file_path.name}: {e}")
                errors += 1
        
        # Rebuild embeddings if any files were processed
        if processed > 0:
            logger.info(f"[CPU Auto-Ingest] 🔧 Rebuilding embeddings after processing {processed} files...")
            if self.rebuild_embeddings():
                logger.info(f"[CPU Auto-Ingest] ✅ Embeddings updated successfully")
            else:
                logger.error(f"[CPU Auto-Ingest] ❌ Failed to rebuild embeddings")
        
        result = {
            "processed": processed,
            "skipped": skipped,
            "errors": errors,
            "total_chunks": len(self.chunks)
        }
        
        logger.info(f"[CPU Auto-Ingest] 📊 Scan complete: {result}")
        return result
    
    def start_watching(self):
        """Start file watching in background thread"""
        if self.watching:
            logger.warning(f"[CPU Auto-Ingest] ⚠️ Already watching")
            return
        
        self.watching = True
        self.watch_thread = threading.Thread(target=self._watch_loop, daemon=True)
        self.watch_thread.start()
        logger.info(f"[CPU Auto-Ingest] 👀 Started file watching")
    
    def stop_watching(self):
        """Stop file watching"""
        self.watching = False
        if self.watch_thread:
            self.watch_thread.join()
        logger.info(f"[CPU Auto-Ingest] ⏹️ Stopped file watching")
    
    def _watch_loop(self):
        """File watching loop"""
        last_scan_time = 0
        
        while self.watching:
            try:
                current_time = time.time()
                
                # Scan every 30 seconds
                if current_time - last_scan_time > 30:
                    result = self.scan_and_process()
                    if result["processed"] > 0:
                        logger.info(f"[CPU Auto-Ingest] 🔄 Auto-processed {result['processed']} files")
                    last_scan_time = current_time
                
                time.sleep(5)  # Check every 5 seconds
                
            except Exception as e:
                logger.error(f"[CPU Auto-Ingest] ❌ Error in watch loop: {e}")
                time.sleep(10)  # Wait longer on error


def main():
    """Main execution for testing"""
    print("\n" + "="*80)
    print("  🔄 CPU FAISS AUTO-INGESTION TEST")
    print("="*80)
    
    # Initialize auto-ingest
    auto_ingest = CPUFAISSAutoIngest()
    
    # Load existing embeddings
    auto_ingest.load_existing_embeddings()
    
    # Scan and process
    result = auto_ingest.scan_and_process()
    
    print(f"\n[CPU Auto-Ingest] 📊 Results:")
    print(f"  - Processed: {result['processed']}")
    print(f"  - Skipped: {result['skipped']}")
    print(f"  - Errors: {result['errors']}")
    print(f"  - Total chunks: {result['total_chunks']}")
    
    print("\n" + "="*80)
    print("  ✅ CPU FAISS AUTO-INGESTION TEST COMPLETE!")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
